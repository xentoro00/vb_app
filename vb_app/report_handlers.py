# apps/vb_app/vb_app/report_handlers.py

import frappe
import json
import io
from frappe import _
from frappe.utils import flt
from frappe.desk.query_report import run as original_run # We import the original to call it later

# Try importing openpyxl, handle if missing
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    pass

# ==============================================================================
# 1. THE SECURE RUN METHOD (Intercepts frappe.desk.query_report.run)
# ==============================================================================
@frappe.whitelist()
@frappe.read_only()
def secure_run_report(report_name, filters=None, user=None, ignore_prepared_report=False, custom_columns=None, is_tree=False, parent_field=None, are_default_filters=True):
    if not user:
        user = frappe.session.user

    # Normalize filters
    if filters and isinstance(filters, str):
        try:
            filters = json.loads(filters)
        except Exception:
            filters = {}
    if not isinstance(filters, dict):
        filters = {}

    # --- SECURITY CHECK START ---
    # Skip for System Managers or Admins
    if user != "Administrator" and "System Manager" not in frappe.get_roles(user):
        user_perms = frappe.permissions.get_user_permissions(user)
        
        if user_perms and user_perms.get("Company"):
            allowed_docs = user_perms.get("Company", [])
            allowed_companies = [d.get("doc") for d in allowed_docs]
            
            requested_company = filters.get("company")

            # CRITICAL: Strict Check
            if requested_company and requested_company not in allowed_companies:
                frappe.throw(f"Security Alert: You do not have permission to view data for Company '{requested_company}'")

            # Force filter if empty or invalid
            if not requested_company or requested_company not in allowed_companies:
                if allowed_companies:
                    filters["company"] = allowed_companies[0]
                else:
                    # If user has NO company permissions at all
                    frappe.throw("Security Alert: No Company Permissions assigned to your user.")
    # --- SECURITY CHECK END ---

    # Call the original Frappe function with the safe filters
    return original_run(
        report_name=report_name,
        filters=filters,
        user=user,
        ignore_prepared_report=ignore_prepared_report,
        custom_columns=custom_columns,
        is_tree=is_tree,
        parent_field=parent_field,
        are_default_filters=are_default_filters
    )


# ==============================================================================
# 2. THE CUSTOM EXCEL EXPORT
# ==============================================================================
@frappe.whitelist(allow_guest=False)
def export_custom_excel(report_name, filters=None):
    # 1. ROBUST FILTER PARSING
    if not filters: filters = {}
    if isinstance(filters, str):
        try:
            filters = json.loads(filters)
        except:
            filters = {}
    
    # 2. RE-RUN SECURITY CHECK (Double safety)
    if "System Manager" not in frappe.get_roles():
        user_perms = frappe.permissions.get_user_permissions(frappe.session.user)
        if user_perms and user_perms.get("Company"):
            allowed_docs = user_perms.get("Company", [])
            allowed_companies = [d.get("doc") for d in allowed_docs]
            requested_company = filters.get("company")
            
            if requested_company and requested_company not in allowed_companies:
                frappe.throw(f"Security Alert: Access denied for '{requested_company}'")
            if not requested_company or requested_company not in allowed_companies:
                if allowed_companies: filters["company"] = allowed_companies[0]

    # 3. GET DATA
    try:
        # We call our own secure run just to be safe/consistent
        report_result = secure_run_report(report_name, filters)
    except Exception as e:
        frappe.msgprint(f"Error running report: {str(e)}")
        return

    data = []
    if isinstance(report_result, dict):
        data = report_result.get("result", [])
    elif isinstance(report_result, (list, tuple)) and len(report_result) >= 2:
        data = report_result[1]
    
    if not data:
        frappe.msgprint("No data found for the selected filters.")
        return

    # 4. CREATE WORKBOOK
    wb = Workbook()
    ws = wb.active
    
    # --- STYLES ---
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    def style(cell, fill=None, font=None, align=None, border_=None):
        if fill: cell.fill = fill
        if font: cell.font = font
        if align: cell.alignment = align
        if border_: cell.border = border_

     # ============================================================
    # Libri i Blerjes
    # ============================================================
    if report_name in ["Purchase Invoice", "Libri i Blerjes", "Purchase Invoice Custom"]:
        ws.title = "Libri i Blerjes - Kuartale"

        # Fonts & Styles
        font_header_narrow = Font(name="Arial Narrow", size=9.5, bold=False)
        font_header_bold_10 = Font(name="Arial Narrow", size=10, bold=True)
        font_data = Font(name="Calibri", size=11, bold=False)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_bottom_nowrap = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
        align_bottom_nowrap_center = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
        align_data = Alignment(horizontal=None, vertical="bottom", wrap_text=False)
        fill_header_blue = PatternFill("solid", fgColor="DAE6F0") 
        fill_sub_blue = PatternFill("solid", fgColor="9EB6CE")

        # Columns
        col_widths = [5, 12, 15, 25, 20] + [15] * 6 + [15] 
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        # Headers
        ws.merge_cells("A1:A3"); ws["A1"] = "Nr."
        ws.merge_cells("B1:C1"); ws["B1"] = "Fatura"
        ws.merge_cells("D1:E1"); ws["D1"] = "Shitësi"
        ws.merge_cells("F1:H1"); ws["F1"] = "Blerje vendore"
        ws.merge_cells("I1:K1"); ws["I1"] = "Importe"
        ws.merge_cells("L1:L2"); ws["L1"] = "Blerjet totale"
        ws.merge_cells("B2:B3"); ws["B2"] = "Data"
        ws.merge_cells("C2:C3"); ws["C2"] = "Numri i faturës"
        ws.merge_cells("D2:D3"); ws["D2"] = "Emri i shitësit"
        ws.merge_cells("E2:E3"); ws["E2"] = "Numri Unik Identifikues / Numri Fiskal / Numri Personal i shitësit"
        ws["F2"] = "Blerje e mallrave / shërbimeve"; ws["G2"] = "Blerje për shpenzime"
        ws["H2"] = "Blerje investive"; ws["I2"] = "Importe të mallrave"
        ws["J2"] = "Importe investive"; ws["K2"] = "Blerje të shërbimeve jashtë vendit"

        row_3_labels = ["a", "b", "c", "d", "e", "f", "g=a+b+c+d+e+f"]
        for idx, label in enumerate(row_3_labels):
            ws.cell(row=3, column=6 + idx).value = label

        # Apply Styles
        for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=12):
            for cell in row: style(cell, fill_header_blue, font_header_narrow, align_center, border)
        for col in range(2, 12): style(ws.cell(row=1, column=col), fill_header_blue, font_header_bold_10, align_bottom_nowrap_center, border)
        for col in range(6, 13): style(ws.cell(row=3, column=col), fill_sub_blue, font_header_narrow, align_bottom_nowrap, border)

        # Data
        start_row = 4
        for i, row in enumerate(data, start=start_row):
            val_date = str(row.get("data") or row.get("posting_date") or "")
            val_id = str(row.get("nr._fatures_furnitorit") or row.get("nr._brendshem") or "")
            val_supplier = str(row.get("furnitori") or "")
            val_tax_id = str(row.get("nr._fiskal") or "")
            val_base = flt(row.get("baza_e_blerjeve", 0))
            val_imports = flt(row.get("import_/_pa_tvsh%", 0))
            val_total = flt(row.get("totale", 0))

            ws.cell(row=i, column=1, value=i - 3)
            ws.cell(row=i, column=2, value=val_date)
            ws.cell(row=i, column=3, value=val_id)
            ws.cell(row=i, column=4, value=val_supplier)
            ws.cell(row=i, column=5, value=val_tax_id)
            c_f = ws.cell(row=i, column=6, value=val_base); c_f.number_format = '#,##0.00'
            c_g = ws.cell(row=i, column=7, value=0.00); c_g.number_format = '#,##0.00'
            c_h = ws.cell(row=i, column=8, value=0.00); c_h.number_format = '#,##0.00'
            c_i = ws.cell(row=i, column=9, value=val_imports); c_i.number_format = '#,##0.00'
            c_j = ws.cell(row=i, column=10, value=0.00); c_j.number_format = '#,##0.00'
            c_k = ws.cell(row=i, column=11, value=0.00); c_k.number_format = '#,##0.00'
            c_l = ws.cell(row=i, column=12, value=val_total); c_l.number_format = '#,##0.00'
            for cell in ws[i]: style(cell, None, font_data, align_data, border)

	# ============================================================
    # LIBRI I BLERJES - 3 MUJOR
    # ============================================================
    elif report_name == "Libri i Blerjes - 3 mujor":
        ws.title = "Libri i Blerjet TVSH"

        # Fonts & Aligns
        font_12_bold = Font(name="Arial Narrow", size=12, bold=True)
        font_11_bold = Font(name="Arial Narrow", size=11, bold=True)
        font_95_norm = Font(name="Arial Narrow", size=9.5, bold=False)
        font_data = Font(name="Calibri", size=11, bold=False)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_center_nowrap = Alignment(horizontal="center", vertical="center", wrap_text=False)
        align_bottom_center = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
        align_bottom_nowrap = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
        align_data = Alignment(horizontal=None, vertical="bottom", wrap_text=False)

        # Fills
        fill_grey_c2 = PatternFill("solid", fgColor="FFFFFF")
        fill_grey_d6 = PatternFill("solid", fgColor="D6D4CA")
        fill_green_e1 = PatternFill("solid", fgColor="E1FFE1")
        fill_white = PatternFill("solid", fgColor="FFFFFF")

        # Columns
        col_widths = [5, 12, 15, 25, 15, 15] + [12] * 23 
        for i, w in enumerate(col_widths, 1):
            col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
            ws.column_dimensions[col_letter].width = w

        # Row 1
        ws.merge_cells("A1:A2"); ws["A1"] = "Nr."
        ws.merge_cells("B1:C1"); ws["B1"] = "Fatura"
        ws.merge_cells("D1:F1"); ws["D1"] = "Shitësi"
        ws.merge_cells("G1:J1"); ws["G1"] = "Blerjet dhe Importet e liruara dhe me TVSH jo të zbritshme"
        ws.merge_cells("K1:S1"); ws["K1"] = "Blerjet dhe Importet e tatushme me 18%, si dhe rregullimet e zbritjeve"
        ws.merge_cells("T1:AB1"); ws["T1"] = "Blerjet dhe Importet e tatushme me 8%, si dhe rregullimet e zbritjeve"
        ws.merge_cells("AC1:AC2"); ws["AC1"] = "Totali i TVSH-së së zbritshme me 18% dhe 8%"

        # Row 2
        ws["B2"] = "Data"; ws["C2"] = "Numri i faturës"; ws["D2"] = "Emri i shitësit"; ws["E2"] = "Numri Fiskal i shitësit"; ws["F2"] = "Numri i TVSH-së së shitësit"
        sub_headers = [("G", "Blerjet dhe importet pa TVSH"), ("H", "Blerjet dhe importet investive pa TVSH"), ("I", "Blerjet dhe importet me TVSH jo të zbritshme"), ("J", "Blerjet dhe importet investive me TVSH jo të zbritshme"), ("K", "Importet"), ("L", "Importet investive"), ("M", "Blerjet vendore"), ("N", "Blerjet investive vendore"), ("O", "Nota debitore e pranuar, nota kreditore e lëshuar"), ("P", "Fatura e borxhit të keq e lëshuar"), ("Q", "Rregullimet për të ulur TVSH-në për pagesë"), ("R", "E drejta e kreditimit të TVSH-së në lidhje me Ngarkesën e Kundërt"), ("S", "Totali i TVSH-së së zbritshme me 18%"), ("T", "Importet"), ("U", "Importet investive"), ("V", "Blerjet vendore"), ("W", "Blerjet investive vendore"), ("X", "Blerjet nga fermerët (aplikimi i normës së sheshtë)"), ("Y", "Nota debitore e pranuar, nota kreditore e lëshuar"), ("Z", "Fatura e borxhit të keq e lëshuar"), ("AA", "Rregullimet për të ulur TVSH-në / Ngarkesa e kundërt për Energjinë"), ("AB", "Totali i TVSH-së së zbritshme me 8%")]
        for col, text in sub_headers: ws[f"{col}2"] = text

        # Row 3
        ws.merge_cells("A3:F3"); ws["A3"] = "Numri i kutisë në Deklaratën e TVSH-së"
        refs = {"G": "[31]", "H": "[32]", "I": "[33]", "J": "[34]", "K": "[35]", "L": "[39]", "M": "[43]", "N": "[47]", "O": "[53]", "P": "[57]", "Q": "[61]", "R": "[65]", "S": "[K1]", "T": "[37]", "U": "[41]", "V": "[45]", "W": "[49]", "X": "[51]", "Y": "[55]", "Z": "[59]", "AA": "[63]", "AB": "[K2]", "AC": "[67]"}
        for col, ref in refs.items(): ws[f"{col}3"] = ref

        # Styling
        for col in range(1, 30):
            cell = ws.cell(row=1, column=col)
            s_fill, s_font, s_align = fill_grey_c2, font_12_bold, align_bottom_center
            if col == 1: s_font = font_95_norm; s_align = align_center
            elif 2 <= col <= 6: s_align = align_bottom_nowrap
            elif 7 <= col <= 28: s_font = font_11_bold; s_align = align_bottom_nowrap
            elif col == 29: s_fill = fill_grey_d6; s_font = font_95_norm; s_align = align_center
            style(cell, s_fill, s_font, s_align, border)

        for col in range(1, 30):
            cell = ws.cell(row=2, column=col)
            s_fill, s_font, s_align = fill_grey_c2, font_95_norm, align_center
            if col == 19 or col == 28: s_fill = fill_green_e1
            elif col == 29: s_fill = fill_grey_d6
            style(cell, s_fill, s_font, s_align, border)

        for col in range(1, 30):
            cell = ws.cell(row=3, column=col)
            style(cell, fill_grey_d6, font_95_norm, align_bottom_nowrap, border)

        # Data
        for i, row in enumerate(data, start=4):
            val_date = str(row.get("data") or row.get("posting_date") or "")
            val_id = str(row.get("nr._fatures_furnitorit") or row.get("nr._brendshem") or "")
            val_supplier = str(row.get("furnitori") or "")
            val_tax_id = str(row.get("nr._fiskal") or "")
            
            ws.cell(row=i, column=1, value=i - 3)
            ws.cell(row=i, column=2, value=val_date)
            ws.cell(row=i, column=3, value=val_id)
            ws.cell(row=i, column=4, value=val_supplier)
            ws.cell(row=i, column=5, value=val_tax_id)
            ws.cell(row=i, column=6, value=val_tax_id)
            for col_idx in range(7, 30): 
                c = ws.cell(row=i, column=col_idx, value=0.00)
                c.number_format = '#,##0.00'
            for c in ws[i]: style(c, None, font_data, align_data, border)


    # ============================================================
    # OPTION B: NEW LIBRI I SHITJES (Wide Template A-Y)
    # ============================================================
    elif report_name in ["Libri i Shitjes", "Sales Invoice"]:
        ws.title = "Libri i Shitjes TVSH"

        # Fonts & Aligns
        font_bold_12 = Font(name="Arial Narrow", size=12, bold=True)
        font_norm_9 = Font(name="Arial Narrow", size=9, bold=False)
        font_norm_9_5 = Font(name="Arial Narrow", size=9.5, bold=False)
        font_data = Font(name="Arial Narrow", size=11)
        
        align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_bottom_center_nowrap = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
        align_center_nowrap = Alignment(horizontal="center", vertical="center", wrap_text=False) # New
        align_data = Alignment(horizontal=None, vertical="bottom", wrap_text=True)

        # Fills
        fill_white = PatternFill("solid", fgColor="FFFFFF")
        fill_k2 = PatternFill("solid", fgColor="EAF0F6") # Light Blue/Grey for R2, W2
        fill_grey_custom = PatternFill("solid", fgColor="D6D4CA") # Grey for X1, X2

        # Column Config (Reduced by 1 column, total 24 -> A to X)
        # G-L (6), M-R (6), S-W (5), X (1) = 18 acc cols + 6 base = 24
        col_widths = [5, 12, 15, 25, 15, 15] + [12] * 18
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'+chr(64+i-26)].width = w

        # Headers Row 1
        ws.merge_cells("A1:A2"); ws["A1"] = "Nr."
        ws.merge_cells("B1:C1"); ws["B1"] = "Fatura"
        ws.merge_cells("D1:F1"); ws["D1"] = "Blerësi"
        ws.merge_cells("G1:L1"); ws["G1"] = "Shitjet e liruara nga TVSH"
        ws.merge_cells("M1:R1"); ws["M1"] = "Shitjet e tatueshme me normën 18%"
        ws.merge_cells("S1:W1"); ws["S1"] = "Shitjet e tatueshme me normën 8%" # Reduced Range
        ws.merge_cells("X1:X2"); ws["X1"] = "Totali i TVSH-së së llogaritur me 18% dhe 8%" # Shifted from Y to X
        
        # Row 2 Sub Headers
        ws["B2"] = "Data"; ws["C2"] = "Numri i faturës"
        ws["D2"] = "Emri i blerësit"; ws["E2"] = "Numri Fiskal i blerësit"
        ws["F2"] = "Numri i TVSH-së së blerësit"

        headers_acc = [
            ("G", "Shitjet e liruara pa të drejtë kreditimi"), ("H", "Shitjet e shërbimeve jashtë vendit"),
            ("I", "Shitjet brenda vendit me ngarkesë të kundërt të TVSH-së"), ("J", "Shitjet tjera të liruara me të drejtë kreditimi"),
            ("K", "Totali i shitjeve të liruara me të drejtë kreditimi"), ("L", "Eksportet"),
            # 18%
            ("M", "Shitjet e tatueshme"), ("N", "Nota debitore e lëshuar, nota kreditore e pranuar"),
            ("O", "Fatura e borxhit të keq e pranuar"), ("P", "Rregullimet për të rritur TVSH-në"),
            ("Q", "Blerjet që i nënshtrohen ngarkesës së kundërt"), ("R", "Totali i TVSH-së së llogaritur me normën 18%"),
            # 8%
            ("S", "Shitjet e tatueshme"), ("T", "Nota debitore e lëshuar, nota kreditore e pranuar"),
            ("U", "Fatura e borxhit të keq e pranuar"), 
            ("V", "Rregullimet për të rritur TVSH-në / Ngarkesa e kundërt për Energjinë me normë 8%"), # Updated
            ("W", "Totali i TVSH-së së llogaritur me normën 8%"), # Moved from X, Text Updated
        ]
        for col, text in headers_acc:
            ws[f"{col}2"] = text
            ws[f"{col}2"].alignment = align_center_wrap

        # Row 3 References
        ws.merge_cells("A3:F3"); ws["A3"] = "Numri i kutisë në Deklaratën e TVSH-së"
        refs = {
            "G": "[9]", "H": "[10a]", "I": "[10b]", "J": "[10c]", 
            "K": "[10] = [10a]+[10b]+[10c]", # Updated
            "L": "[11]",
            "M": "[12]", "N": "[16]", "O": "[20]", "P": "[24]", "Q": "[28]", "R": "[K1]",
            "S": "[14]", "T": "[18]", "U": "[22]", "V": "[26]", 
            "W": "[K2]", 
            "X": "[30]" 
        }
        for col, ref in refs.items(): ws[f"{col}3"] = ref

        # --- STYLING APPLICATION ---
        
        # General Row Styles
        for cell in ws["1:1"]: style(cell, fill_white, font_bold_12, align_bottom_center_nowrap, border)
        for cell in ws["2:2"]: style(cell, fill_white, font_norm_9, align_center_wrap, border)
        for cell in ws["3:3"]: style(cell, fill_grey_custom, font_norm_9_5, align_bottom_center_nowrap, border)
        
        # 1. A1, A2: No wrap text, middle align
        style(ws["A1"], fill_white, font_norm_9, align_center_nowrap, border)
        style(ws["A2"], fill_white, font_norm_9, align_center_nowrap, border) # Merged, but safe to style

        # 2. B2: No wrap text
        ws["B2"].alignment = align_center_nowrap
        ws["D2"].alignment = align_center_nowrap

        # 3. K2: Blue Fill
        style(ws["K2"], fill_k2, font_norm_9, align_center_wrap, border)
        
        # 4. R2, W2: Background #EAF0F6, Font size 9
        style(ws["R2"], fill_k2, font_norm_9, align_center_wrap, border)
        style(ws["W2"], fill_k2, font_norm_9, align_center_wrap, border)

        # 5. X1, X2: Background #D6D4CA, not bold, 9.5 font size, middle align
        style(ws["X1"], fill_grey_custom, font_norm_9_5, align_center_wrap, border)
        style(ws["X2"], fill_grey_custom, font_norm_9_5, align_center_wrap, border)
        
        style(ws["P3"], fill_grey_custom, font_norm_9_5, align_center_wrap, border)
        # Data Rows
        start_row = 4
        for i, row in enumerate(data, start=start_row):
            val_date = str(row.get("data") or row.get("posting_date") or "")
            val_id = str(row.get("nr._fatures") or row.get("name") or "")
            val_customer = str(row.get("klienti") or row.get("customer_name") or "")
            val_tax_id = str(row.get("nr._fiskal") or row.get("tax_id") or "")
            
            ws.cell(row=i, column=1, value=i - 3)
            ws.cell(row=i, column=2, value=val_date)
            ws.cell(row=i, column=3, value=val_id)
            ws.cell(row=i, column=4, value=val_customer)
            ws.cell(row=i, column=5, value=val_tax_id)
            ws.cell(row=i, column=6, value=val_tax_id)

            # Accounting columns (G to X) - Note range is now up to 25
            for col_idx in range(7, 25): 
                c = ws.cell(row=i, column=col_idx, value=0.00)
                c.number_format = '#,##0.00'

            for c in ws[i]: style(c, None, font_data, align_data, border)


    # ============================================================
    # OPTION C: OLD LIBRI I SHITJES - 3 MUJOR (Orange Template)
    # ============================================================
    elif report_name in ["Libri i Shitjes - 3 mujor", "Sales Invoice Custom Report", "Libri i Shitjes - 1 mujor"]:
        ws.title = "Libri i Shitjes Kuartale"
        header_fill = PatternFill("solid", fgColor="FDE9D9")
        subheader_fill = PatternFill("solid", fgColor="FABF8F")
        header_font = Font(name="Arial Narrow", size=9.5, bold=False)
        font_header_exception = Font(name="Arial Narrow", size=11, bold=True)
        data_font = Font(name="Arial Narrow", size=10)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_align = Alignment(horizontal=None, vertical="bottom", wrap_text=True)

        widths = [5, 12, 18, 25, 25, 20, 20, 12, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.row_dimensions[1].height = 25; ws.row_dimensions[2].height = 60; ws.row_dimensions[3].height = 20

        ws.merge_cells("B1:C1"); ws.merge_cells("D1:E1"); ws.merge_cells("A1:A3")
        ws.merge_cells("B2:B3"); ws.merge_cells("C2:C3"); ws.merge_cells("D2:D3"); ws.merge_cells("E2:E3")
        ws.merge_cells("F1:F2"); ws.merge_cells("G1:G2"); ws.merge_cells("H1:H2"); ws.merge_cells("I1:I2")

        ws["A1"] = "Nr."; ws["B1"] = "Fatura"; ws["D1"] = "Blerësi"
        ws["F1"] = "Shitje e mallrave / shërbimeve brenda vendit"
        ws["G1"] = "Shitje e shërbimeve jashtë vendit"
        ws["H1"] = "Eksporte"; ws["I1"] = "Shitjet totale"
        ws["B2"] = "Data"; ws["C2"] = "Numri i faturës"
        ws["D2"] = "Emri i blerësit"; ws["E2"] = "Numri Unik / Identifikues / Numri Fiskal i blerësit / Company Tax ID"
        ws["F3"] = "a"; ws["G3"] = "b"; ws["H3"] = "c"; ws["I3"] = "d=a+b+c"

        for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=9):
            for cell in row: style(cell, header_fill, header_font, align_center, border)
        for cell_coord in ["F3", "G3", "H3", "I3"]:
            style(ws[cell_coord], subheader_fill, header_font, align_center, border)
        style(ws["B1"], font=font_header_exception)
        style(ws["D1"], font=font_header_exception)

        start_row = 4
        for i, row in enumerate(data, start=start_row):
            print("row")
            print(row)
            val_date = str(row.get("data") or row.get("posting_date") or "")
            val_id = str(row.get("nr._fatures") or row.get("name") or "")
            val_customer = str(row.get("klienti") or row.get("customer_name") or "")
            val_tax_id = str(row.get("nr._fiskal") or row.get("tax_id") or "")
            val_base = flt(row.get("qarkullimi_i_tatueshem_(baza)", 0))
            val_export = flt(row.get("eksport_/_0%", 0))
            val_total = flt(row.get("totale", 0))

            ws.cell(row=i, column=1, value=i - 3)
            ws.cell(row=i, column=2, value=val_date)
            ws.cell(row=i, column=3, value=val_id)
            ws.cell(row=i, column=4, value=val_customer)
            ws.cell(row=i, column=5, value=val_tax_id)
            c_f = ws.cell(row=i, column=6, value=val_base); c_f.number_format = '#,##0.00'
            c_g = ws.cell(row=i, column=7, value=val_export); c_g.number_format = '#,##0.00'
            c_h = ws.cell(row=i, column=8, value=0); c_h.number_format = '#,##0.00'
            c_i = ws.cell(row=i, column=9, value=val_total); c_i.number_format = '#,##0.00'
            for cell in ws[i]: style(cell, None, data_font, data_align, border)

    # ============================================================
    # OPTION D: NO TEMPLATE MATCH
    # ============================================================
    else:
        print(report_name)
        frappe.msgprint("No custom template defined for this report.")
        return

    # 5. SAVE AND RETURN
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = f"{report_name}_custom.xlsx"
    frappe.response['filecontent'] = output.read()
    frappe.response['type'] = 'binary'