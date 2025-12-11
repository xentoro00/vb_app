import frappe
import zipfile
from io import BytesIO
from frappe import _
from frappe.utils import flt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Side, Border, Font, Alignment

# Import the original builder functions for the "else" cases
from frappe.core.doctype.data_import.exporter import build_csv_response, build_xlsx_response

# 1. THE MAIN FUNCTION
def custom_build_response(self):
    # --- SALES INVOICE LOGIC ---
    if self.doctype == "Sales Invoice" and self.file_type == "Excel":
        try:
            original_headers = self.csv_array[0]
            # 1. ROBUST ID FINDER
            # Sometimes the header is "ID", sometimes "Name", sometimes "Series"
            if "ID" in original_headers:
                id_idx = original_headers.index("ID")
            elif "Name" in original_headers:
                id_idx = original_headers.index("Name")
            elif "Series" in original_headers:
                id_idx = original_headers.index("Series")
            else:
                # Fallback: The ID is almost always the first column (Index 0) or second (Index 1)
                # if the user hasn't shuffled them. But safe to throw error if we can't find it.
                frappe.throw("The <b>ID</b> column is missing from the export. Please make sure to check 'ID' (or Name) in the column selection.")
            date_idx = original_headers.index("Date")
            id_idx = original_headers.index("ID")
            customer_idx = original_headers.index("Customer")
            tax_id_idx = original_headers.index("Company Tax ID")
            total_qty_idx = original_headers.index("Total Quantity")
        except ValueError as e:
            frappe.throw(f"Missing required field in export: {e}. Please ensure Date, ID, Customer, Company Tax ID, and Total Quantity are selected.")
            return

        # Generate File 1: Fatura
        fatura_file = self.generate_libri_i_shitjes_kuartale_excel(original_headers)
        
        # Generate File 2: Libri i Shitjes
        libri_file = self.generate_libri_i_shitjes_tvsh_excel(original_headers)

        # Zip them together
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr("Libri i Shitjes Kuartale.xlsx", fatura_file.getvalue())
            zip_file.writestr("Libri i Shitjes TVSH.xlsx", libri_file.getvalue())

        zip_buffer.seek(0)
        frappe.response.filename = "Libri_I_Shitjes.zip"
        frappe.response.filecontent = zip_buffer.read()
        frappe.response.type = "binary"
        return

    # --- PURCHASE INVOICE LOGIC ---
    elif self.doctype == "Purchase Invoice" and self.file_type == "Excel":
        try:
            original_headers = self.csv_array[0]
            date_idx = original_headers.index("Date")
            id_idx = original_headers.index("ID")
            supplier_idx = original_headers.index("Supplier")
            tax_id_idx = original_headers.index("Tax Id")
        except ValueError as e:
             frappe.throw(f"Missing required field in export: {e}. Please ensure Date, ID, Supplier, and Tax ID are selected.")
             return

        # Generate File 1: Libri i Blerjes Kuartale
        libri_blerjes_kuartale = self.generate_libri_blerjes_kuartale_excel(original_headers)
        
        # Generate File 2: Libri i Blerjes TVSH
        libri_blerjes_tvsh = self.generate_libri_i_blerjes_tvsh_excel(original_headers)

        # Zip them together
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr("Libri i Blerjes Kuartale.xlsx", libri_blerjes_kuartale.getvalue())
            zip_file.writestr("Libri i Blerjes TVSH.xlsx", libri_blerjes_tvsh.getvalue())

        zip_buffer.seek(0)
        frappe.response.filename = "Libri_I_Blerjes.zip"
        frappe.response.filecontent = zip_buffer.read()
        frappe.response.type = "binary"
        return

    # --- DEFAULT LOGIC (Original Frappe Behavior) ---
    elif self.file_type == "CSV":
        build_csv_response(self.get_csv_array_for_export(), _(self.doctype))
    elif self.file_type == "Excel":
        build_xlsx_response(self.get_csv_array_for_export(), _(self.doctype))

# 2. PASTE YOUR 4 HELPER FUNCTIONS BELOW HERE
# (I have referenced them above, so just paste the def generate_... parts here exactly as you wrote them)

# =================================================================================
# TEMPLATE 1: Libri i Shitjes Kuartale
# =================================================================================
def generate_libri_i_shitjes_kuartale_excel(self, original_headers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Libri i Shitjes Kuartale"

    # Indices
    date_idx = original_headers.index("Date")
    id_idx = original_headers.index("ID")
    customer_idx = original_headers.index("Customer")
    tax_id_idx = original_headers.index("Company Tax ID")
    total_qty_idx = original_headers.index("Total Quantity")

    # Styles
    header_fill = PatternFill("solid", fgColor="FDE9D9")
    subheader_fill = PatternFill("solid", fgColor="FABF8F")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial Narrow", size=9.5, bold=False)
    font_header_exception = Font(name="Arial Narrow", size=11, bold=True)
    data_font = Font(name="Arial Narrow", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal=None, vertical="bottom", wrap_text=True)

    def style(cell, fill=None, font=None, align=None, border_=None):
        if fill: cell.fill = fill
        if font: cell.font = font
        if align: cell.alignment = align
        if border_: cell.border = border_

    # Dimensions
    widths = [5, 12, 18, 25, 25, 20, 20, 12, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 60
    ws.row_dimensions[3].height = 20

    # Merges
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:E1")
    ws.merge_cells("A1:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    ws.merge_cells("D2:D3")
    ws.merge_cells("E2:E3")
    ws.merge_cells("F1:F2")
    ws.merge_cells("G1:G2")
    ws.merge_cells("H1:H2")
    ws.merge_cells("I1:I2")

    # Content
    ws["A1"] = "Nr."
    ws["B1"] = "Fatura"
    ws["D1"] = "Blerësi"
    ws["F1"] = "Shitje e mallrave / shërbimeve brenda vendit"
    ws["G1"] = "Shitje e shërbimeve jashtë vendit"
    ws["H1"] = "Eksporte"
    ws["I1"] = "Shitjet totale"
    ws["B2"] = "Data"
    ws["C2"] = "Numri i faturës"
    ws["D2"] = "Emri i blerësit"
    ws["E2"] = "Numri Unik / Identifikues / Numri Fiskal i blerësit / Company Tax ID"
    ws["F3"] = "a"
    ws["G3"] = "b"
    ws["H3"] = "c"
    ws["I3"] = "d=a+b+c"

    # Apply Styles
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=9):
        for cell in row:
            style(cell, header_fill, header_font, align_center, border)
    
    for cell_coord in ["F3", "G3", "H3", "I3"]:
        style(ws[cell_coord], subheader_fill, header_font, align_center, border)
    
    style(ws["B1"], font=font_header_exception)
    style(ws["D1"], font=font_header_exception)

    # Data
    start_row = 4
    for i, data_row in enumerate(self.csv_array[1:], start=start_row):
        total_f = flt(data_row[total_qty_idx])
        total_g = 0
        total_h = 0
        total_i = total_f + total_g + total_h

        ws.cell(row=i, column=1, value=i - 3)
        ws.cell(row=i, column=2, value=data_row[date_idx])
        ws.cell(row=i, column=3, value=data_row[id_idx])
        ws.cell(row=i, column=4, value=data_row[customer_idx])
        ws.cell(row=i, column=5, value=data_row[tax_id_idx])

        cell_f = ws.cell(row=i, column=6, value=total_f)
        cell_f.number_format = '#,##0.00'
        cell_g = ws.cell(row=i, column=7, value=total_g)
        cell_g.number_format = '#,##0.00'
        cell_h = ws.cell(row=i, column=8, value=total_h)
        cell_h.number_format = '#,##0.00'
        cell_i = ws.cell(row=i, column=9, value=total_i)
        cell_i.number_format = '#,##0.00'

        for cell in ws[i]:
            style(cell, None, data_font, data_align, border)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# =================================================================================
# TEMPLATE 2: Libri i Shitjes TVSH
# =================================================================================
def generate_libri_i_shitjes_tvsh_excel(self, original_headers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Libri i Shitjes TVSH"

    # Indices (Reuse from Fatura)
    date_idx = original_headers.index("Date")
    id_idx = original_headers.index("ID")
    customer_idx = original_headers.index("Customer")
    tax_id_idx = original_headers.index("Company Tax ID")

    # Styles
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # New Fonts
    header_font_bold_12 = Font(name="Arial Narrow", size=12, bold=True)
    header_font_norm = Font(name="Arial Narrow", size=9, bold=False)
    header_font_9_5_nobold = Font(name="Arial Narrow", size=9.5, bold=False)
    data_font_11 = Font(name="Arial Narrow", size=11)
    
    # New Aligns
    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_bottom_center_wrap = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    data_align = Alignment(horizontal=None, vertical="bottom", wrap_text=True)

    # Fills
    fill_grey = PatternFill("solid", fgColor="D9D9D9")
    fill_light_blue = PatternFill("solid", fgColor="DCE6F1")
    fill_white = PatternFill("solid", fgColor="FFFFFF")
    fill_k2 = PatternFill("solid", fgColor="EAF0F6")
    fill_grey_custom = PatternFill("solid", fgColor="D6D4CA")

    def style(cell, fill=None, font=None, align=None, border_=None):
        if fill: cell.fill = fill
        if font: cell.font = font
        if align: cell.alignment = align
        if border_: cell.border = border_

    # === Column Config ===
    # A: Nr, B: Data, C: Numri, D: Emri, E: N.Fiskal, F: N.TVSH, G-Y Accounting
    col_widths = [5, 12, 15, 25, 15, 15] + [12] * 20 
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'+chr(64+i-26)].width = w

    # === Headers Row 1 (Merges & Text) ===
    # A merged A1:A2 (Nr)
    ws.merge_cells("A1:A2")
    ws["A1"] = "Nr."

    # Fatura (B1:C1)
    ws.merge_cells("B1:C1")
    ws["B1"] = "Fatura"

    # Blerësi (D1:F1)
    ws.merge_cells("D1:F1")
    ws["D1"] = "Blerësi"

    # Shitjet e liruara (G1:L1)
    ws.merge_cells("G1:L1")
    ws["G1"] = "Shitjet e liruara nga TVSH"

    # Eksportet (M1)
    # 18% Group: M, N, O, P, Q, R (Total 18%)
    ws.merge_cells("M1:R1")
    ws["M1"] = "Shitjet e tatueshme me normën 18%"

    # 8% Group: S, T, U, V, W, X (Total 8%)
    ws.merge_cells("S1:X1")
    ws["S1"] = "Shitjet e tatueshme me normën 8%"

    # Total Final: Y1:Y2
    ws.merge_cells("Y1:Y2")
    ws["Y1"] = "Totali i TVSH-së së llogaritur me 18% dhe 8%"
    
    # === Row 2 (Sub Headers) ===
    ws["B2"] = "Data"
    ws["C2"] = "Numri i faturës"
    ws["D2"] = "Emri i blerësit"
    ws["E2"] = "Numri Fiskal i blerësit"
    ws["F2"] = "Numri i TVSH-së së blerësit"

    headers_acc = [
        ("G", "Shitjet e liruara pa të drejtë kreditimi"),
        ("H", "Shitjet e shërbimeve jashtë vendit"),
        ("I", "Shitjet brenda vendit me ngarkesë të kundërt të TVSH-së"),
        ("J", "Shitjet tjera të liruara me të drejtë kreditimi"),
        ("K", "Totali i shitjeve të liruara me të drejtë kreditimi"),
        ("L", "Eksportet"),
        # 18%
        ("M", "Shitjet e tatueshme"),
        ("N", "Nota debitore e lëshuar, nota kreditore e pranuar"),
        ("O", "Fatura e borxhit të keq e pranuar"),
        ("P", "Rregullimet për të rritur TVSH-në"),
        ("Q", "Blerjet që i nënshtrohen ngarkesës së kundërt"),
        ("R", "Totali i TVSH-së së llogaritur me normën 18%"),
        # 8%
        ("S", "Shitjet e tatueshme"),
        ("T", "Nota debitore e lëshuar, nota kreditore e pranuar"),
        ("U", "Fatura e borxhit të keq e pranuar"),
        ("V", "Rregullimet për të rritur TVSH-në"),
        ("W", "Blerjet... Energjinë..."),
        ("X", "Totali i TVSH-së së llogaritur me normën 8%"),
    ]

    for col, text in headers_acc:
        ws[f"{col}2"] = text
        ws[f"{col}2"].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # === Row 3 (References) ===
    # Merge A3:F3 -> "Numri i kutisë në Deklaratën e TVSH-së"
    ws.merge_cells("A3:F3")
    ws["A3"] = "Numri i kutisë në Deklaratën e TVSH-së"

    refs = {
        "G": "[9]", "H": "[10a]", "I": "[10b]", "J": "[10c]", 
        "K": "[10] = [10a]+[10b]+[10c]", "L": "[11]",
        "M": "[12]", "N": "[16]", "O": "[20]", "P": "[24]", "Q": "[28]", "R": "[K1]",
        "S": "[14]", "T": "[18]", "U": "[22]", "V": "[26]", "W": "[K2]", "X": "[30]"
    }
    for col, ref in refs.items():
        ws[f"{col}3"] = ref

    # === Styling Headers ===
    # Top Row (Bold 12, Bottom Align)
    for cell in ws["1:1"]:
        style(cell, fill_white, header_font_bold_12, align_bottom_center_wrap, border)
    
    # Row 2 (Normal 9)
    for cell in ws["2:2"]:
        style(cell, fill_white, header_font_norm, align_center_wrap, border)

    # Row 3 (Grey fill custom, Font 9.5 No Bold, Bottom Align)
    for cell in ws["3:3"]:
        style(cell, fill_grey_custom, header_font_9_5_nobold, align_bottom_center_wrap, border)
    
    # === Exceptions ===
    
    # A1:A2 -> Font 9, No Bold
    style(ws["A1"], fill_white, header_font_norm, align_bottom_center_wrap, border)

    # D2 -> No Wrap
    ws["D2"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # K2 -> Fill Custom Blue
    style(ws["K2"], fill_k2, header_font_norm, align_center_wrap, border)

    # X1, X2 -> Fill Grey Custom, Font 9.5 No Bold
    # Note: X1 is part of merge S1:X1, styling S1 would affect all S-X. Styling X2 works.
    style(ws["X2"], fill_grey_custom, header_font_9_5_nobold, align_center_wrap, border)
    

    # === Data Rows ===
    start_row = 4
    for i, data_row in enumerate(self.csv_array[1:], start=start_row):
        # Mapping
        ws.cell(row=i, column=1, value=i - 3) # Nr
        ws.cell(row=i, column=2, value=data_row[date_idx]) # Date
        ws.cell(row=i, column=3, value=data_row[id_idx]) # Invoice ID
        ws.cell(row=i, column=4, value=data_row[customer_idx]) # Customer Name
        ws.cell(row=i, column=5, value=data_row[tax_id_idx]) # Tax ID
        ws.cell(row=i, column=6, value=data_row[tax_id_idx]) # VAT ID (Using Tax ID again as common in KS)

        # Accounting columns (G to Y) - initializing to 0.00
        for col_idx in range(7, 26): # G is 7, Y is 25
            c = ws.cell(row=i, column=col_idx, value=0.00)
            c.number_format = '#,##0.00'

        for cell in ws[i]:
            # Data Rows -> Font 11
            style(cell, None, data_font_11, data_align, border)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# =================================================================================
# TEMPLATE 3: Libri i Blerjes - Kuartale
# =================================================================================
def generate_libri_blerjes_kuartale_excel(self, original_headers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Libri i Blerjes - Kuartale"

    # Indices
    date_idx = original_headers.index("Date")
    id_idx = original_headers.index("ID")
    supplier_idx = original_headers.index("Supplier")
    tax_id_idx = original_headers.index("Tax Id")

    # Styles
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # New Fonts
    font_header_narrow = Font(name="Arial Narrow", size=9.5, bold=False)
    font_header_bold_10 = Font(name="Arial Narrow", size=10, bold=True)
    font_data = Font(name="Calibri", size=11, bold=False)
    
    # Alignment
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_bottom_nowrap = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
    align_bottom_nowrap_center = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
    align_data = Alignment(horizontal=None, vertical="bottom", wrap_text=False) # No horizontal, No Wrap

    # Fills
    fill_header_blue = PatternFill("solid", fgColor="DAE6F0") 
    fill_sub_blue = PatternFill("solid", fgColor="9EB6CE")

    def style(cell, fill=None, font=None, align=None, border_=None):
        if fill: cell.fill = fill
        if font: cell.font = font
        if align: cell.alignment = align
        if border_: cell.border = border_

    # === Column Config ===
    col_widths = [5, 12, 15, 25, 20] + [15] * 6 + [15] 
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # === Header Row 1 ===
    # A1 merged with A3 (A1:A3)
    ws.merge_cells("A1:A3")
    ws["A1"] = "Nr."

    ws.merge_cells("B1:C1")
    ws["B1"] = "Fatura"

    ws.merge_cells("D1:E1")
    ws["D1"] = "Shitësi"

    ws.merge_cells("F1:H1")
    ws["F1"] = "Blerje vendore"

    ws.merge_cells("I1:K1")
    ws["I1"] = "Importe"

    # L1 merged with L2 (per snippet)
    ws.merge_cells("L1:L2")
    ws["L1"] = "Blerjet totale"

    # === Header Row 2 (Sub-headers & Merges with Row 3) ===
    # B2:B3
    ws.merge_cells("B2:B3")
    ws["B2"] = "Data"
    
    # C2:C3
    ws.merge_cells("C2:C3")
    ws["C2"] = "Numri i faturës"
    
    # D2:D3
    ws.merge_cells("D2:D3")
    ws["D2"] = "Emri i shitësit"
    
    # E2:E3
    ws.merge_cells("E2:E3")
    ws["E2"] = "Numri Unik Identifikues / Numri Fiskal / Numri Personal i shitësit"
    
    ws["F2"] = "Blerje e mallrave / shërbimeve"
    ws["G2"] = "Blerje për shpenzime"
    ws["H2"] = "Blerje investive"
    ws["I2"] = "Importe të mallrave"
    ws["J2"] = "Importe investive"
    ws["K2"] = "Blerje të shërbimeve jashtë vendit"

    # === Header Row 3 (Labels F-L) ===
    # Note: B,C,D,E merged down to 3. F-L are free.
    row_3_labels = ["a", "b", "c", "d", "e", "f", "g=a+b+c+d+e+f"]
    # Cols F to L -> Indices 6 to 12
    for idx, label in enumerate(row_3_labels):
        col = 6 + idx # F=6
        cell = ws.cell(row=3, column=col)
        cell.value = label

    # === Styling ===
    # All headers (Row 1, 2, 3) get Arial Narrow 9.5 No Bold
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=12):
        for cell in row:
            style(cell, fill_header_blue, font_header_narrow, align_center, border)

    # Exceptions for Row 1 B-K (Indices 2-11)
    # "BCDEFGHIJK:1 should be bold, font size of 10 bottom aligned an dno wrap text"
    for col in range(2, 12): # B=2, K=11. range stops before 12.
        cell = ws.cell(row=1, column=col)
        style(cell, fill_header_blue, font_header_bold_10, align_bottom_nowrap_center, border)

    # Exceptions for Row 3 F-L (Indices 6-12)
    # "the 3:FGHIJKL cells should not have wrap text and should be bottom aligned"
    # Using fill_sub_blue for them per snippet
    for col in range(6, 13): 
        cell = ws.cell(row=3, column=col)
        style(cell, fill_sub_blue, font_header_narrow, align_bottom_nowrap, border)

    # === Data Rows ===
    start_row = 4
    for i, data_row in enumerate(self.csv_array[1:], start=start_row):
        ws.cell(row=i, column=1, value=i - 3) 
        ws.cell(row=i, column=2, value=data_row[date_idx]) 
        ws.cell(row=i, column=3, value=data_row[id_idx]) 
        ws.cell(row=i, column=4, value=data_row[supplier_idx]) 
        ws.cell(row=i, column=5, value=data_row[tax_id_idx]) 

        # Accounting columns (F to L)
        for col_idx in range(6, 13): 
            c = ws.cell(row=i, column=col_idx, value=0.00)
            c.number_format = '#,##0.00'
        
        for cell in ws[i]:
            style(cell, None, font_data, align_data, border)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# =================================================================================
# TEMPLATE 4: Libri i Blerjet  TVSH
# =================================================================================
def generate_libri_i_blerjes_tvsh_excel(self, original_headers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Libri i Blerjet  TVSH"

    # Indices
    date_idx = original_headers.index("Date")
    id_idx = original_headers.index("ID")
    supplier_idx = original_headers.index("Supplier")
    tax_id_idx = original_headers.index("Tax Id")

    # Styles
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Fonts
    font_header_12_bold = Font(name="Arial Narrow", size=12, bold=True)
    font_header_11_bold = Font(name="Arial Narrow", size=11, bold=True)
    font_header_95_norm = Font(name="Arial Narrow", size=9.5, bold=False)
    font_data = Font(name="Calibri", size=11, bold=False)
    
    # Alignment
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_center_nowrap = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_bottom_center = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    align_bottom_center_nowrap = Alignment(horizontal="center", vertical="bottom", wrap_text=False)
    align_data = Alignment(horizontal=None, vertical="bottom", wrap_text=False)

    


    # Fills
    fill_grey_c2 = PatternFill("solid", fgColor="C2C6D0") # Grey for Headers
    fill_grey_d6 = PatternFill("solid", fgColor="D6D4CA") # Grey for Row 3 & AC
    fill_green_e1 = PatternFill("solid", fgColor="E1FFE1") # Green for S2, AB2
    fill_white = PatternFill("solid", fgColor="FFFFFF")

    def style(cell, fill=None, font=None, align=None, border_=None):
        if fill: cell.fill = fill
        if font: cell.font = font
        if align: cell.alignment = align
        if border_: cell.border = border_

    # === Column Config (A to AC = 29 cols) ===
    # A: Nr, B: Data, C: Numri, D: Emri, E: N.Fiskal, F: N.TVSH, G-AC Accounting
    col_widths = [5, 12, 15, 25, 15, 15] + [12] * 23 
    for i, w in enumerate(col_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w

    # === Header Row 1 ===
    ws.merge_cells("A1:A2")
    ws["A1"] = "Nr."